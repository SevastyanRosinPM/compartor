# coding: utf-8
"""
main.py
Запуск без аргументов: ищет Mos.csv и Invaders.csv рядом с этим файлом и создаёт report.html
Поддерживает:
 - объединение спринтов по номеру (игнорируя даты)
 - двустороннее сопоставление (META-XXX в title и наоборот)
 - HTML визуал: колонки по спринтам -> ДИТ / Invaders, подсветка карточек
 - фильтр по задаче
 - фильтр по спринту
 - выгрузку в Excel
 - разделение на свимлайны "Задачи" и "Баги"
 - вывод статусов задач
Запуск: нажать Run в IDE (PyCharm/VSCode и т.д.)
Зависимости: pandas, openpyxl
    pip install pandas openpyxl
"""

import re
import html
from pathlib import Path
import pandas as pd
from datetime import datetime

# -------------------------
# Настройки
# -------------------------
MOS_NAME = "Mos.csv"
INV_NAME = "Invaders.csv"
OUT_NAME = "report.html"
EXCEL_NAME = "comparison_report.xlsx"

# Базовые URL для задач
MOS_BASE_URL = "https://itpm.mos.ru/browse/"
INV_BASE_URL = "https://jira.theinvaders.ru/browse/"

# Префиксы для задач Invaders
INV_PREFIXES = ['MT-', 'PART-', 'FEATURE-', 'BUG-', 'TASK-', 'EPIC-', 'STORY-', 'IMPROVEMENT-']

# -------------------------
# Вспомогательные функции
# -------------------------
def read_csv_guess(path: Path) -> pd.DataFrame:
    try:
        return pd.read_csv(path, encoding="utf-8-sig")
    except Exception:
        # fallback
        return pd.read_csv(path, encoding="cp1251", errors="ignore")

def canonical_sprint(s: str) -> str:
    """Вернуть 'Спринт N' по любой строке, содержащей 'Спринт' и номер.
       Если номер не найден — 'Нет спринта'"""
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return "Нет спринта"
    s = str(s)
    m = re.search(r'Спринт\s*(\d+)', s, flags=re.IGNORECASE)
    if m:
        return f"Спринт {int(m.group(1))}"
    # иногда в данных может быть 'META Спринт 13'
    m2 = re.search(r'META\s*Спринт\s*(\d+)', s, flags=re.IGNORECASE)
    if m2:
        return f"Спринт {int(m2.group(1))}"
    return "Нет спринта"

def extract_meta_key_from_text(s: str):
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return None
    m = re.search(r'(META-\d+)', str(s), flags=re.IGNORECASE)
    return m.group(1).upper() if m else None

def extract_inv_key_from_text(s: str):
    """Извлечение ключа Invaders (MT-, PART-, FEATURE- и т.д.) из текста"""
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return None
    
    s_str = str(s).upper()
    
    # Проверяем все возможные префиксы Invaders
    for prefix in INV_PREFIXES:
        pattern = fr'({prefix}\d+)'
        m = re.search(pattern, s_str, re.IGNORECASE)
        if m:
            return m.group(1).upper()
    
    return None

def get_task_url(task_id: str, task_type: str) -> str:
    """Получить URL для задачи по её ID и типу"""
    if not task_id or pd.isna(task_id):
        return "#"
    
    task_id_str = str(task_id).strip().upper()
    
    if task_type == 'mos':
        # Для ДИТ: добавляем META- если нет
        if task_id_str.startswith('META-'):
            return f"{MOS_BASE_URL}{task_id_str}"
        else:
            # Проверяем, есть ли номер в ID
            num_match = re.search(r'(\d+)', task_id_str)
            if num_match:
                return f"{MOS_BASE_URL}META-{num_match.group(1)}"
            return "#"
    elif task_type == 'inv':
        # Для Invaders: проверяем все возможные префиксы
        for prefix in INV_PREFIXES:
            if task_id_str.startswith(prefix):
                return f"{INV_BASE_URL}{task_id_str}"
        
        # Если префикса нет, пробуем извлечь номер и добавляем MT- по умолчанию
        num_match = re.search(r'(\d+)', task_id_str)
        if num_match:
            # Но сначала проверим, может быть в самом ID есть префикс, но в нижнем регистре
            lower_id = str(task_id).lower()
            for prefix in INV_PREFIXES:
                prefix_lower = prefix.lower()
                if prefix_lower in lower_id:
                    # Извлекаем полный ключ с префиксом
                    pattern = fr'({prefix_lower}\d+)'
                    m = re.search(pattern, lower_id)
                    if m:
                        return f"{INV_BASE_URL}{m.group(1).upper()}"
            
            # Если не нашли префикс, используем MT- по умолчанию
            return f"{INV_BASE_URL}MT-{num_match.group(1)}"
        return "#"
    return "#"

def normalize_inv_key(key_str: str):
    """Нормализация ключа Invaders - извлекаем правильный формат"""
    if not key_str or pd.isna(key_str):
        return None
    
    # Пробуем извлечь ключ с любым из префиксов
    key_upper = str(key_str).upper()
    
    for prefix in INV_PREFIXES:
        if key_upper.startswith(prefix):
            # Убедимся, что после префикса есть номер
            if re.search(r'\d+$', key_upper[len(prefix):]):
                return key_upper
    
    # Если не нашли префикс, пробуем извлечь из строки
    for prefix in INV_PREFIXES:
        pattern = fr'({prefix}\d+)'
        m = re.search(pattern, key_upper)
        if m:
            return m.group(1)
    
    return None

def find_status_column(df, system_name):
    """Найти колонку со статусом в DataFrame"""
    status_columns = []
    
    # Возможные названия колонок со статусом
    possible_names = [
        'Статус', 'Status', 'Состояние', 'State',
        'Статус задачи', 'Status of the task',
        'Статус проблемы', 'Issue Status'
    ]
    
    for col in df.columns:
        col_str = str(col).strip()
        
        # Проверяем точное совпадение
        if col_str in possible_names:
            return col
        
        # Проверяем частичное совпадение
        col_lower = col_str.lower()
        if any(name.lower() in col_lower for name in ['статус', 'status', 'состояние', 'state']):
            status_columns.append(col)
    
    # Если нашли несколько подходящих, выбираем первую
    if status_columns:
        return status_columns[0]
    
    # Если не нашли, проверяем содержимое колонок
    for col in df.columns:
        # Берем первые несколько непустых значений
        sample_values = df[col].dropna().head(5)
        if len(sample_values) > 0:
            # Проверяем, содержат ли значения типичные статусы
            status_keywords = ['открыт', 'в работе', 'готово', 'закрыт', 'отложен',
                              'open', 'in progress', 'done', 'closed', 'resolved']
            for val in sample_values:
                val_str = str(val).lower()
                if any(keyword in val_str for keyword in status_keywords):
                    return col
    
    print(f"  ⚠️ Для {system_name} не найдена колонка со статусом. Доступные колонки: {list(df.columns)[:10]}...")
    return None

# -------------------------
# Сопоставление
# -------------------------
def match_two_way(mos_df, inv_df):
    """
    Возвращаем:
      matches: список кортежей (mos_index, inv_index)
      mos_used: set индексов
      inv_used: set индексов
    Алгоритм:
      1) прямое совпадение по ключу 'Ключ проблемы' (равенство)
      2) если у Mos есть ключ META-XXX и он встречается в теме Invaders -> match
      3) если у Inv есть ключ META-XXX и он встречается в теме Mos -> match
    """
    matches = []
    mos_used = set()
    inv_used = set()

    # Убедимся, что колонки существуют
    if 'Ключ проблемы' not in mos_df.columns:
        mos_df['Ключ проблемы'] = mos_df.index.astype(str)
    if 'Ключ проблемы' not in inv_df.columns:
        inv_df['Ключ проблемы'] = None

    # 1) прямое совпадение ключей (case-insensitive)
    inv_key_map = {}
    for ji, j in inv_df.iterrows():
        v = j.get('Ключ проблемы')
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        v_str = str(v).upper()
        inv_key_map[v_str] = ji

    for mi, m in mos_df.iterrows():
        mk = m.get('Ключ проблемы')
        if mk is None or (isinstance(mk, float) and pd.isna(mk)):
            continue
        mk_u = str(mk).upper()
        if mk_u in inv_key_map:
            ji = inv_key_map[mk_u]
            matches.append((mi, ji))
            mos_used.add(mi)
            inv_used.add(ji)

    # 2) ключ Mos в теме Invaders
    for mi, m in mos_df.iterrows():
        if mi in mos_used:
            continue
        mk = m.get('Ключ проблемы')
        if mk is None or (isinstance(mk, float) and pd.isna(mk)):
            continue
        mk_u = str(mk).upper()
        for ji, j in inv_df.iterrows():
            if ji in inv_used:
                continue
            title = str(j.get('Тема') or j.get('title') or "")
            if mk_u in title.upper():
                matches.append((mi, ji))
                mos_used.add(mi)
                inv_used.add(ji)
                break

    # 3) ключ Inv in topic Mos
    for ji, j in inv_df.iterrows():
        if ji in inv_used:
            continue
        jk = j.get('Ключ проблемы')
        if jk is None or (isinstance(jk, float) and pd.isna(jk)):
            continue
        jk_u = str(jk).upper()
        for mi, m in mos_df.iterrows():
            if mi in mos_used:
                continue
            title = str(m.get('Тема') or "")
            if jk_u in title.upper():
                matches.append((mi, ji))
                mos_used.add(mi)
                inv_used.add(ji)
                break

    return matches, mos_used, inv_used

def categorize_and_prepare(mos_df, inv_df, matches, mos_used, inv_used):
    """
    Возвращает структуру categorized:
      {
        'match': [ {mos_id, inv_id, mos_title, inv_title, mos_sprint, inv_sprint, mos_url, inv_url, is_bug, mos_status, inv_status}, ... ],
        'diff_sprint': [...],
        'mos_only': [...],
        'inv_only': [...]
      }
    """
    categorized = {'match': [], 'diff_sprint': [], 'mos_only': [], 'inv_only': []}

    # Находим колонки со статусами
    mos_status_col = find_status_column(mos_df, "ДИТ")
    inv_status_col = find_status_column(inv_df, "Invaders")
    
    if mos_status_col:
        print(f"  ✓ Найдена колонка статуса для ДИТ: '{mos_status_col}'")
    else:
        print(f"  ✗ Колонка статуса для ДИТ не найдена")
    
    if inv_status_col:
        print(f"  ✓ Найдена колонка статуса для Invaders: '{inv_status_col}'")
    else:
        print(f"  ✗ Колонка статуса для Invaders не найдена")

    for mi, ji in matches:
        m = mos_df.loc[mi]
        j = inv_df.loc[ji]
        ms = canonical_sprint(m.get('Компоненты') if 'Компоненты' in m.index else m.get('sprint'))
        js = canonical_sprint(j.get('Пользовательское поле (Релизный спринт)') if 'Пользовательское поле (Релизный спринт)' in j.index else j.get('sprint'))
        
        mos_id = m.get('Ключ проблемы')
        inv_id = j.get('Ключ проблемы')
        
        # Нормализуем ключ Invaders
        if inv_id:
            normalized_inv_id = normalize_inv_key(inv_id)
            if normalized_inv_id:
                inv_id = normalized_inv_id
        
        # Получаем URL для задач
        mos_url = get_task_url(mos_id, 'mos')
        inv_url = get_task_url(inv_id, 'inv')
        
        # Получаем статусы задач
        mos_status = m.get(mos_status_col) if mos_status_col else None
        inv_status = j.get(inv_status_col) if inv_status_col else None
        
        # Очищаем статусы от NaN
        if mos_status and isinstance(mos_status, float) and pd.isna(mos_status):
            mos_status = None
        if inv_status and isinstance(inv_status, float) and pd.isna(inv_status):
            inv_status = None
        
        # Получаем названия задач
        mos_title = m.get('Тема') or ""
        inv_title = j.get('Тема') or j.get('title') or ""
        
        # Определяем, является ли задача багом
        # Проверяем в названии Invaders наличие [Баг]
        is_bug = False
        if isinstance(inv_title, str) and '[Баг]' in inv_title:
            is_bug = True
        elif isinstance(mos_title, str) and '[Баг]' in mos_title:
            is_bug = True
        
        rec = {
            'mos_id': mos_id,
            'inv_id': inv_id,
            'mos_title': mos_title,
            'inv_title': inv_title,
            'mos_sprint': ms,
            'inv_sprint': js,
            'mos_url': mos_url,
            'inv_url': inv_url,
            'is_bug': is_bug,
            'mos_status': str(mos_status) if mos_status else "Неизвестно",
            'inv_status': str(inv_status) if inv_status else "Неизвестно"
        }
        if ms == js:
            categorized['match'].append(rec)
        else:
            categorized['diff_sprint'].append(rec)

    # mos only
    for mi, m in mos_df.iterrows():
        if mi in mos_used:
            continue
        ms = canonical_sprint(m.get('Компоненты'))
        mos_id = m.get('Ключ проблемы')
        mos_url = get_task_url(mos_id, 'mos')
        
        # Получаем статус задачи
        mos_status = m.get(mos_status_col) if mos_status_col else None
        if mos_status and isinstance(mos_status, float) and pd.isna(mos_status):
            mos_status = None
        
        # Определяем, является ли багом
        mos_title = m.get('Тема') or ""
        is_bug = isinstance(mos_title, str) and '[Баг]' in mos_title
        
        categorized['mos_only'].append({
            'mos_id': mos_id,
            'mos_title': mos_title,
            'mos_sprint': ms,
            'mos_url': mos_url,
            'is_bug': is_bug,
            'mos_status': str(mos_status) if mos_status else "Неизвестно"
        })

    # inv only
    for ji, j in inv_df.iterrows():
        if ji in inv_used:
            continue
        js = canonical_sprint(j.get('Пользовательское поле (Релизный спринт)') if 'Пользовательское поле (Релизный спринт)' in j.index else j.get('sprint'))
        inv_id = j.get('Ключ проблемы')
        
        # Нормализуем ключ Invaders
        if inv_id:
            normalized_inv_id = normalize_inv_key(inv_id)
            if normalized_inv_id:
                inv_id = normalized_inv_id
        
        inv_url = get_task_url(inv_id, 'inv')
        inv_title = j.get('Тема') or j.get('title') or ""
        
        # Получаем статус задачи
        inv_status = j.get(inv_status_col) if inv_status_col else None
        if inv_status and isinstance(inv_status, float) and pd.isna(inv_status):
            inv_status = None
        
        # Определяем, является ли багом
        is_bug = isinstance(inv_title, str) and '[Баг]' in inv_title
        
        categorized['inv_only'].append({
            'inv_id': inv_id,
            'inv_title': inv_title,
            'inv_sprint': js,
            'inv_url': inv_url,
            'is_bug': is_bug,
            'inv_status': str(inv_status) if inv_status else "Неизвестно"
        })

    return categorized

def export_to_excel(categorized, out_file: Path, mos_df, inv_df):
    """
    Создает Excel файл с несколькими листами:
    1. Сводка (статистика)
    2. Совпадения
    3. Разные спринты
    4. Только ДИТ
    5. Только Invaders
    6. Исходные данные ДИТ
    7. Исходные данные Invaders
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    print(f"Создание Excel файла: {out_file}")
    
    # Создаем новую книгу
    wb = Workbook()
    
    # Удаляем дефолтный лист
    if 'Sheet' in wb.sheetnames:
        ws = wb['Sheet']
        wb.remove(ws)
    
    # Настройки стилей
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0F1724", end_color="0F1724", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Лист 1: Сводка
    ws_summary = wb.create_sheet("Сводка")
    
    # Заголовок
    ws_summary.merge_cells('A1:D1')
    ws_summary['A1'] = "Сводный отчет по сопоставлению задач ДИТ и Invaders"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary['A1'].alignment = center_alignment
    
    ws_summary['A3'] = "Дата создания отчета:"
    ws_summary['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Статистика
    ws_summary['A5'] = "Статистика"
    ws_summary['A5'].font = Font(bold=True, size=12)
    
    # Подсчет багов
    bug_count_match = sum(1 for item in categorized['match'] if item.get('is_bug', False))
    bug_count_diff = sum(1 for item in categorized['diff_sprint'] if item.get('is_bug', False))
    bug_count_mos = sum(1 for item in categorized['mos_only'] if item.get('is_bug', False))
    bug_count_inv = sum(1 for item in categorized['inv_only'] if item.get('is_bug', False))
    total_bugs = bug_count_match + bug_count_diff + bug_count_mos + bug_count_inv
    
    # Статистика по статусам
    status_counts = {}
    for cat in categorized.values():
        for item in cat:
            for sys in ['mos', 'inv']:
                status_key = f'{sys}_status'
                if status_key in item:
                    status = item[status_key]
                    if status not in status_counts:
                        status_counts[status] = 0
                    status_counts[status] += 1
    
    stats_data = [
        ["Показатель", "Количество"],
        ["Всего задач ДИТ", len(categorized['match']) + len(categorized['diff_sprint']) + len(categorized['mos_only'])],
        ["Всего задач Invaders", len(categorized['match']) + len(categorized['diff_sprint']) + len(categorized['inv_only'])],
        ["Совпадения (одинаковые спринты)", len(categorized['match'])],
        ["Совпадения (разные спринты)", len(categorized['diff_sprint'])],
        ["Только в ДИТ", len(categorized['mos_only'])],
        ["Только в Invaders", len(categorized['inv_only'])],
        ["Всего совпадений", len(categorized['match']) + len(categorized['diff_sprint'])],
        ["Процент совпадений", f"{(len(categorized['match']) + len(categorized['diff_sprint'])) / max(len(categorized['match']) + len(categorized['diff_sprint']) + len(categorized['mos_only']), 1) * 100:.1f}%"],
        ["Всего багов", total_bugs],
        ["Баги в совпадениях", bug_count_match + bug_count_diff],
        ["Баги только в ДИТ", bug_count_mos],
        ["Баги только в Invaders", bug_count_inv]
    ]
    
    for i, row in enumerate(stats_data):
        for j, value in enumerate(row):
            cell = ws_summary.cell(row=i+6, column=j+1)
            cell.value = value
            cell.border = border_style
            if i == 0:  # Заголовок таблицы
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
    
    # Добавляем статистику по статусам
    row_offset = len(stats_data) + 8
    ws_summary.cell(row=row_offset, column=1, value="Распределение по статусам").font = Font(bold=True, size=12)
    
    status_data = [["Статус", "Количество"]]
    for status, count in sorted(status_counts.items()):
        status_data.append([status, count])
    
    for i, row in enumerate(status_data):
        for j, value in enumerate(row):
            cell = ws_summary.cell(row=row_offset + i + 1, column=j+1)
            cell.value = value
            cell.border = border_style
            if i == 0:  # Заголовок таблицы
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
    
    # Лист 2: Совпадения
    ws_matches = wb.create_sheet("Совпадения")
    matches_headers = ["Спринт", "Ключ ДИТ", "Название ДИТ", "Статус ДИТ", "Ссылка ДИТ", 
                      "Ключ Invaders", "Название Invaders", "Статус Invaders", "Ссылка Invaders", "Статус", "Тип"]
    
    for col, header in enumerate(matches_headers, 1):
        cell = ws_matches.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border_style
    
    row = 2
    for item in categorized['match']:
        ws_matches.cell(row=row, column=1, value=item['mos_sprint']).border = border_style
        ws_matches.cell(row=row, column=2, value=item['mos_id']).border = border_style
        ws_matches.cell(row=row, column=3, value=item['mos_title']).border = border_style
        ws_matches.cell(row=row, column=4, value=item.get('mos_status', 'Неизвестно')).border = border_style
        ws_matches.cell(row=row, column=5, value=item['mos_url']).border = border_style
        ws_matches.cell(row=row, column=6, value=item['inv_id']).border = border_style
        ws_matches.cell(row=row, column=7, value=item['inv_title']).border = border_style
        ws_matches.cell(row=row, column=8, value=item.get('inv_status', 'Неизвестно')).border = border_style
        ws_matches.cell(row=row, column=9, value=item['inv_url']).border = border_style
        ws_matches.cell(row=row, column=10, value="Совпадение").border = border_style
        ws_matches.cell(row=row, column=11, value="Баг" if item.get('is_bug', False) else "Задача").border = border_style
        row += 1
    
    # Лист 3: Разные спринты
    ws_diff = wb.create_sheet("Разные спринты")
    diff_headers = ["Спринт ДИТ", "Спринт Invaders", "Ключ ДИТ", "Название ДИТ", "Статус ДИТ", 
                   "Ссылка ДИТ", "Ключ Invaders", "Название Invaders", "Статус Invaders", "Ссылка Invaders", "Статус", "Тип"]
    
    for col, header in enumerate(diff_headers, 1):
        cell = ws_diff.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border_style
    
    row = 2
    for item in categorized['diff_sprint']:
        ws_diff.cell(row=row, column=1, value=item['mos_sprint']).border = border_style
        ws_diff.cell(row=row, column=2, value=item['inv_sprint']).border = border_style
        ws_diff.cell(row=row, column=3, value=item['mos_id']).border = border_style
        ws_diff.cell(row=row, column=4, value=item['mos_title']).border = border_style
        ws_diff.cell(row=row, column=5, value=item.get('mos_status', 'Неизвестно')).border = border_style
        ws_diff.cell(row=row, column=6, value=item['mos_url']).border = border_style
        ws_diff.cell(row=row, column=7, value=item['inv_id']).border = border_style
        ws_diff.cell(row=row, column=8, value=item['inv_title']).border = border_style
        ws_diff.cell(row=row, column=9, value=item.get('inv_status', 'Неизвестно')).border = border_style
        ws_diff.cell(row=row, column=10, value=item['inv_url']).border = border_style
        ws_diff.cell(row=row, column=11, value="Разные спринты").border = border_style
        ws_diff.cell(row=row, column=12, value="Баг" if item.get('is_bug', False) else "Задача").border = border_style
        row += 1
    
    # Лист 4: Только ДИТ - ОБНОВЛЕНО: добавлена колонка статуса
    ws_mos_only = wb.create_sheet("Только ДИТ")
    mos_only_headers = ["Спринт", "Ключ ДИТ", "Название ДИТ", "Статус ДИТ", "Ссылка ДИТ", "Тип"]
    
    for col, header in enumerate(mos_only_headers, 1):
        cell = ws_mos_only.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border_style
    
    row = 2
    for item in categorized['mos_only']:
        ws_mos_only.cell(row=row, column=1, value=item['mos_sprint']).border = border_style
        ws_mos_only.cell(row=row, column=2, value=item['mos_id']).border = border_style
        ws_mos_only.cell(row=row, column=3, value=item['mos_title']).border = border_style
        ws_mos_only.cell(row=row, column=4, value=item.get('mos_status', 'Неизвестно')).border = border_style
        ws_mos_only.cell(row=row, column=5, value=item['mos_url']).border = border_style
        ws_mos_only.cell(row=row, column=6, value="Баг" if item.get('is_bug', False) else "Задача").border = border_style
        row += 1
    
    # Лист 5: Только Invaders
    ws_inv_only = wb.create_sheet("Только Invaders")
    inv_only_headers = ["Спринт", "Ключ Invaders", "Название Invaders", "Статус Invaders", "Ссылка Invaders", "Тип"]
    
    for col, header in enumerate(inv_only_headers, 1):
        cell = ws_inv_only.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border_style
    
    row = 2
    for item in categorized['inv_only']:
        ws_inv_only.cell(row=row, column=1, value=item['inv_sprint']).border = border_style
        ws_inv_only.cell(row=row, column=2, value=item['inv_id']).border = border_style
        ws_inv_only.cell(row=row, column=3, value=item['inv_title']).border = border_style
        ws_inv_only.cell(row=row, column=4, value=item.get('inv_status', 'Неизвестно')).border = border_style
        ws_inv_only.cell(row=row, column=5, value=item['inv_url']).border = border_style
        ws_inv_only.cell(row=row, column=6, value="Баг" if item.get('is_bug', False) else "Задача").border = border_style
        row += 1
    
    # Лист 6: Исходные данные ДИТ (ограничим количество колонок)
    ws_mos_raw = wb.create_sheet("Исходные данные ДИТ")
    
    # Выбираем только строковые колонки для избежания ошибок сортировки
    mos_columns = []
    for col in mos_df.columns:
        # Проверяем, что колонка содержит строковые данные
        try:
            # Пробуем взять первую непустую ячейку
            sample_value = mos_df[col].dropna().iloc[0] if not mos_df[col].dropna().empty else ""
            # Если это строка или число, добавляем колонку
            if isinstance(sample_value, (str, int, float)):
                mos_columns.append(col)
        except:
            continue
    
    # Если не нашли подходящих колонок, берем первые 8
    if len(mos_columns) == 0:
        mos_columns = list(mos_df.columns)[:8]
    
    # Ограничиваем количество колонок
    mos_columns = mos_columns[:8]
    
    for col_idx, header in enumerate(mos_columns, 1):
        cell = ws_mos_raw.cell(row=1, column=col_idx)
        cell.value = str(header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border_style
    
    for i, row_data in mos_df.iterrows():
        for j, col_name in enumerate(mos_columns, 1):
            cell = ws_mos_raw.cell(row=i+2, column=j)
            value = row_data[col_name] if col_name in row_data and not pd.isna(row_data[col_name]) else ""
            cell.value = str(value) if not isinstance(value, (int, float)) else value
            cell.border = border_style
    
    # Лист 7: Исходные данные Invaders (ограничим количество колонок)
    ws_inv_raw = wb.create_sheet("Исходные данные Invaders")
    
    # Выбираем только строковые колонки для избежания ошибок сортировки
    inv_columns = []
    for col in inv_df.columns:
        # Проверяем, что колонка содержит строковые данные
        try:
            # Пробуем взять первую непустую ячейку
            sample_value = inv_df[col].dropna().iloc[0] if not inv_df[col].dropna().empty else ""
            # Если это строка или число, добавляем колонку
            if isinstance(sample_value, (str, int, float)):
                inv_columns.append(col)
        except:
            continue
    
    # Если не нашли подходящих колонок, берем первые 8
    if len(inv_columns) == 0:
        inv_columns = list(inv_df.columns)[:8]
    
    # Ограничиваем количество колонок
    inv_columns = inv_columns[:8]
    
    for col_idx, header in enumerate(inv_columns, 1):
        cell = ws_inv_raw.cell(row=1, column=col_idx)
        cell.value = str(header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border_style
    
    for i, row_data in inv_df.iterrows():
        for j, col_name in enumerate(inv_columns, 1):
            cell = ws_inv_raw.cell(row=i+2, column=j)
            value = row_data[col_name] if col_name in row_data and not pd.isna(row_data[col_name]) else ""
            cell.value = str(value) if not isinstance(value, (int, float)) else value
            cell.border = border_style
    
    # Настройка ширины колонок (без сортировки)
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Автонастройка ширины для колонки "Статус ДИТ"
    if 'Только ДИТ' in wb.sheetnames:
        ws_mos_only = wb['Только ДИТ']
        # Настраиваем ширину для колонки статуса (столбец D)
        ws_mos_only.column_dimensions['D'].width = 30  # Ширина для статусов типа "На анализе у исполнителя"
    
    # Сохраняем файл
    wb.save(out_file)
    print(f"Excel файл успешно создан: {out_file}")

# -------------------------
# HTML генерация с разделением на свимлайны и статусами
# -------------------------
def generate_html(categorized, out_file: Path, mos_df, inv_df):
    # собрать все спринты и отсортировать по номеру
    sprint_set = set()
    for cat in categorized.values():
        for it in cat:
            for k, v in it.items():
                if 'sprint' in k and v:
                    sprint_set.add(v)
    # гарантируем 'Нет спринта' если пусто
    if not sprint_set:
        sprint_set.add("Нет спринта")

    def sprint_key(s):
        m = re.search(r'(\d+)', s)
        return int(m.group(1)) if m else 9999

    sorted_sprints = sorted(list(sprint_set), key=sprint_key)

    # Подсчет статистики
    total_tasks = 0
    total_bugs = 0
    status_colors = {
        'готово': '#2f9e44',
        'закрыт': '#2f9e44',
        'выполнено': '#2f9e44',
        'done': '#2f9e44',
        'closed': '#2f9e44',
        'resolved': '#2f9e44',
        'в работе': '#e6b000',
        'в прогрессе': '#e6b000',
        'открыт': '#1e6fe0',
        'новая': '#1e6fe0',
        'to do': '#1e6fe0',
        'open': '#1e6fe0',
        'отложен': '#6b7280',
        'отклонен': '#dc2626',
        'rejected': '#dc2626'
    }
    
    for cat in categorized.values():
        for it in cat:
            total_tasks += 1
            if it.get('is_bug', False):
                total_bugs += 1
    total_regular = total_tasks - total_bugs

    # CSS + JS (приближённый к твоему образцу)
    css = """
    <style>
    body{font-family:Inter, Arial, sans-serif;background:#f6f7fb;margin:0;padding:24px}
    .container{max-width:1300px;margin:0 auto;background:#fff;padding:18px;border-radius:8px;box-shadow:0 6px 18px rgba(20,20,50,0.06)}
    h1{font-size:18px;margin:0 0 12px}
    h2{font-size:16px;margin:20px 0 12px;color:#0f1724}
    .legend{margin-bottom:12px;font-size:13px}
    table{width:100%;border-collapse:collapse}
    th{background:#0f1724;color:#fff;padding:8px;font-weight:600}
    td{vertical-align:top;border:1px solid #e6e9ef;padding:8px;height:320px;overflow:auto}
    .col-head{font-size:13px;background:#fff;padding:6px 8px;border-bottom:1px solid #e6e9ef}
    .task{border-radius:6px;padding:8px;margin-bottom:8px;font-size:13px;box-shadow:0 1px 0 rgba(0,0,0,0.03)}
    .task .id{font-weight:700;font-size:12px;color:#0b4470}
    .task .status{display:inline-block;padding:2px 6px;border-radius:4px;font-size:11px;font-weight:600;margin-left:8px;vertical-align:middle}
    .task .title{color:#0b1724;margin-top:4px}
    .task .title a{color:#0b1724;text-decoration:none;border-bottom:1px dotted #0b4470;}
    .task .title a:hover{color:#0b4470;border-bottom:1px solid #0b4470;}
    .match{background:#e6f6ea;border-left:4px solid #2f9e44}
    .diff{background:#fff8e0;border-left:4px solid #e6b000}
    .mos-only{background:#ffe9e9;border-left:4px solid #d64545}
    .inv-only{background:#e8f1ff;border-left:4px solid #1e6fe0}
    .controls{margin-bottom:12px}
    .btn{display:inline-block;padding:6px 10px;border-radius:6px;background:#0f1724;color:#fff;text-decoration:none;margin-right:8px;font-size:13px;cursor:pointer}
    .export-btn{display:inline-block;padding:6px 10px;border-radius:6px;background:#2f9e44;color:#fff;text-decoration:none;margin-right:8px;font-size:13px;cursor:pointer}
    .export-btn:hover{background:#2b8c3f}
    .small{font-size:12px;color:#6b7280;margin-left:10px}
    .filter-container{margin-top:8px;display:flex;align-items:center;gap:8px}
    .filter-input{padding:6px 10px;border:1px solid #e6e9ef;border-radius:6px;font-size:13px;flex:1;max-width:300px}
    .filter-select{padding:6px 10px;border:1px solid #e6e9ef;border-radius:6px;font-size:13px;max-width:200px}
    .filter-btn{padding:6px 10px;border-radius:6px;background:#0f1724;color:#fff;border:none;cursor:pointer;font-size:13px}
    .filter-btn:hover{background:#1a2638}
    .filter-clear{padding:6px 10px;border-radius:6px;background:#e6e9ef;color:#0f1724;border:none;cursor:pointer;font-size:13px}
    .filter-clear:hover{background:#d1d5db}
    .filter-row{margin-top:8px;display:flex;align-items:center;gap:8px}
    .filter-label{font-size:13px;color:#0f1724;margin-right:4px}
    .task-hidden{display:none}
    .sprint-hidden{display:none}
    .table-container{overflow-x:auto;margin-top:16px}
    .export-section{margin-top:16px;padding:12px;background:#f8fafc;border-radius:6px;border:1px solid #e6e9ef}
    .export-info{font-size:13px;color:#4b5563;margin-top:8px}
    .swimlane{margin-bottom:24px;border:1px solid #e6e9ef;border-radius:8px;overflow:hidden}
    .swimlane-header{background:#f8fafc;padding:12px 16px;border-bottom:1px solid #e6e9ef;cursor:pointer;display:flex;justify-content:space-between;align-items:center}
    .swimlane-header:hover{background:#f1f5f9}
    .swimlane-title{font-weight:600;font-size:14px}
    .swimlane-count{background:#e6e9ef;color:#0f1724;padding:2px 8px;border-radius:12px;font-size:12px}
    .swimlane-content{padding:0}
    .swimlane-collapsed .swimlane-content{display:none}
    .bug-indicator{display:inline-block;background:#fef2f2;color:#dc2626;padding:2px 6px;border-radius:4px;font-size:11px;margin-left:6px;font-weight:600}
    .bug-task{border-left-color:#dc2626 !important}
    .status-ready{background:#2f9e44;color:white}
    .status-inprogress{background:#e6b000;color:white}
    .status-open{background:#1e6fe0;color:white}
    .status-other{background:#6b7280;color:white}
    .status-rejected{background:#dc2626;color:white}
    </style>
    """

    js = """
    <script>
    let originalTableHTML = {tasks: '', bugs: ''};
    let allSprints = [];
    let currentFilteredSprint = '';
    
    document.addEventListener('DOMContentLoaded', function() {
        // Сохраняем оригинальные таблицы и список спринтов
        const tasksTable = document.querySelector('#tasks-swinlane table');
        const bugsTable = document.querySelector('#bugs-swinlane table');
        if (tasksTable) originalTableHTML.tasks = tasksTable.outerHTML;
        if (bugsTable) originalTableHTML.bugs = bugsTable.outerHTML;
        
        const sprintSelect = document.getElementById('sprintFilter');
        allSprints = Array.from(sprintSelect.options).map(opt => opt.value).filter(val => val);
        
        // Инициализируем свимлайны
        initSwimlanes();
    });
    
    function initSwimlanes() {
        const swimlaneHeaders = document.querySelectorAll('.swimlane-header');
        swimlaneHeaders.forEach(header => {
            header.addEventListener('click', function() {
                const swimlane = this.parentElement;
                swimlane.classList.toggle('swimlane-collapsed');
            });
        });
    }
    
    function toggleSwimlane(swimlaneId) {
        const swimlane = document.getElementById(swimlaneId);
        if (swimlane) {
            swimlane.classList.toggle('swimlane-collapsed');
        }
    }
    
    function expandAllSwimlanes() {
        document.querySelectorAll('.swimlane').forEach(swimlane => {
            swimlane.classList.remove('swimlane-collapsed');
        });
    }
    
    function collapseAllSwimlanes() {
        document.querySelectorAll('.swimlane').forEach(swimlane => {
            swimlane.classList.add('swimlane-collapsed');
        });
    }
    
    function toggleClass(cls){
        const els = document.querySelectorAll('.' + cls);
        els.forEach(e => {
            e.style.display = (e.style.display === 'none')? '' : 'none';
        });
    }
    
    function showAll(){
        // Восстанавливаем оригинальные таблицы
        const tasksSwimlane = document.getElementById('tasks-swinlane');
        const bugsSwimlane = document.getElementById('bugs-swinlane');
        
        if (tasksSwimlane && originalTableHTML.tasks) {
            const tableContainer = tasksSwimlane.querySelector('.table-container');
            if (tableContainer) {
                tableContainer.innerHTML = originalTableHTML.tasks;
            }
        }
        
        if (bugsSwimlane && originalTableHTML.bugs) {
            const tableContainer = bugsSwimlane.querySelector('.table-container');
            if (tableContainer) {
                tableContainer.innerHTML = originalTableHTML.bugs;
            }
        }
        
        currentFilteredSprint = '';
        
        // Сбрасываем выпадающий список
        const sprintSelect = document.getElementById('sprintFilter');
        if (sprintSelect) {
            sprintSelect.value = '';
        }
        
        // Показываем все задачи
        ['match','diff','mos-only','inv-only'].forEach(c => {
            document.querySelectorAll('.' + c).forEach(e => e.style.display = '');
        });
    }
    
    function hideAll(){
        ['match','diff','mos-only','inv-only'].forEach(c => {
            document.querySelectorAll('.' + c).forEach(e => e.style.display = 'none');
        });
    }
    
    // Функция фильтрации по задаче
    function filterByTask() {
        const filterInput = document.getElementById('taskFilter');
        const filterValue = filterInput.value.trim().toUpperCase();
        
        if (!filterValue) {
            // Если поле пустое, показываем все задачи
            document.querySelectorAll('.task').forEach(task => {
                task.classList.remove('task-hidden');
            });
            return;
        }
        
        // Скрываем все задачи
        document.querySelectorAll('.task').forEach(task => {
            task.classList.add('task-hidden');
        });
        
        // Показываем только задачи, содержащие фильтр
        document.querySelectorAll('.task').forEach(task => {
            const idElement = task.querySelector('.id');
            const titleElement = task.querySelector('.title');
            
            if (idElement && idElement.textContent.toUpperCase().includes(filterValue)) {
                task.classList.remove('task-hidden');
            } else if (titleElement && titleElement.textContent.toUpperCase().includes(filterValue)) {
                task.classList.remove('task-hidden');
            }
        });
    }
    
    // Функция фильтрации по спринту
    function filterBySprint() {
        const sprintSelect = document.getElementById('sprintFilter');
        const selectedSprint = sprintSelect.value;
        
        if (!selectedSprint || selectedSprint === currentFilteredSprint) {
            return;
        }
        
        currentFilteredSprint = selectedSprint;
        
        // Фильтруем обе таблицы (задачи и баги)
        ['tasks-swinlane', 'bugs-swinlane'].forEach(swimlaneId => {
            const swimlane = document.getElementById(swimlaneId);
            if (!swimlane) return;
            
            const tableContainer = swimlane.querySelector('.table-container');
            const table = swimlane.querySelector('table');
            
            if (!table) {
                console.error('Table not found in', swimlaneId);
                return;
            }
            
            // Получаем оригинальную таблицу для этого свимлайна
            const originalHTML = swimlaneId === 'tasks-swinlane' ? originalTableHTML.tasks : originalTableHTML.bugs;
            
            // Создаем новую таблицу только с выбранным спринтом
            const newTable = document.createElement('table');
            
            // Создаем заголовки для выбранного спринта
            const thead = document.createElement('thead');
            
            // Первая строка заголовков
            const headerRow1 = document.createElement('tr');
            const headerCell1 = document.createElement('th');
            headerCell1.colSpan = 2;
            headerCell1.textContent = selectedSprint;
            headerRow1.appendChild(headerCell1);
            thead.appendChild(headerRow1);
            
            // Вторая строка заголовков
            const headerRow2 = document.createElement('tr');
            const headerCellDit = document.createElement('th');
            headerCellDit.className = 'col-head';
            headerCellDit.textContent = 'ДИТ';
            headerRow2.appendChild(headerCellDit);
            
            const headerCellInv = document.createElement('th');
            headerCellInv.className = 'col-head';
            headerCellInv.textContent = 'Invaders';
            headerRow2.appendChild(headerCellInv);
            thead.appendChild(headerRow2);
            
            newTable.appendChild(thead);
            
            // Создаем тело таблицы
            const tbody = document.createElement('tbody');
            const bodyRow = document.createElement('tr');
            
            // Колонка ДИТ
            const ditCell = document.createElement('td');
            
            // Находим индекс спринта в оригинальной таблице
            const originalTable = document.createElement('div');
            originalTable.innerHTML = originalHTML;
            const originalSprintHeaders = originalTable.querySelectorAll('th[colspan="2"]');
            let sprintIndex = -1;
            
            for (let i = 0; i < originalSprintHeaders.length; i++) {
                if (originalSprintHeaders[i].textContent.trim() === selectedSprint) {
                    sprintIndex = i;
                    break;
                }
            }
            
            if (sprintIndex !== -1) {
                // Получаем все задачи из оригинальной таблицы для этого спринта
                const originalTds = originalTable.querySelectorAll('td');
                const ditCellIndex = sprintIndex * 2;
                
                if (ditCellIndex < originalTds.length) {
                    // Копируем содержимое колонки ДИТ
                    ditCell.innerHTML = originalTds[ditCellIndex].innerHTML;
                }
            }
            
            bodyRow.appendChild(ditCell);
            
            // Колонка Invaders
            const invCell = document.createElement('td');
            
            if (sprintIndex !== -1) {
                const originalTds = originalTable.querySelectorAll('td');
                const invCellIndex = sprintIndex * 2 + 1;
                
                if (invCellIndex < originalTds.length) {
                    // Копируем содержимое колонки Invaders
                    invCell.innerHTML = originalTds[invCellIndex].innerHTML;
                }
            }
            
            bodyRow.appendChild(invCell);
            tbody.appendChild(bodyRow);
            newTable.appendChild(tbody);
            
            // Заменяем таблицу
            tableContainer.innerHTML = '';
            tableContainer.appendChild(newTable);
        });
        
        // Применяем текущий фильтр по задаче, если он есть
        const filterInput = document.getElementById('taskFilter');
        if (filterInput && filterInput.value.trim()) {
            filterByTask();
        }
    }
    
    // Очистка фильтра по спринту
    function clearSprintFilter() {
        const sprintSelect = document.getElementById('sprintFilter');
        sprintSelect.value = '';
        currentFilteredSprint = '';
        
        // Восстанавливаем оригинальные таблицы
        showAll();
        
        // Применяем текущий фильтр по задаче, если он есть
        const filterInput = document.getElementById('taskFilter');
        if (filterInput && filterInput.value.trim()) {
            filterByTask();
        }
    }
    
    // Очистка фильтра по задаче
    function clearFilter() {
        const filterInput = document.getElementById('taskFilter');
        filterInput.value = '';
        
        if (currentFilteredSprint) {
            // Если есть фильтр по спринту, пересоздаем таблицу
            const sprintSelect = document.getElementById('sprintFilter');
            sprintSelect.value = currentFilteredSprint;
            filterBySprint();
        } else {
            // Иначе показываем все задачи
            document.querySelectorAll('.task').forEach(task => {
                task.classList.remove('task-hidden');
            });
        }
    }
    
    // Очистка всех фильтров
    function clearAllFilters() {
        clearFilter();
        clearSprintFilter();
    }
    
    // Обработка нажатия Enter в поле фильтра
    function handleFilterKeyPress(event) {
        if (event.key === 'Enter') {
            filterByTask();
        }
    }
    </script>
    """

    # Функция для получения CSS класса статуса
    def get_status_class(status_str):
        if not status_str:
            return 'status-other'
        
        status_lower = status_str.lower()
        
        if any(word in status_lower for word in ['готово', 'закрыт', 'выполнено', 'done', 'closed', 'resolved']):
            return 'status-ready'
        elif any(word in status_lower for word in ['в работе', 'в прогрессе', 'in progress', 'progress']):
            return 'status-inprogress'
        elif any(word in status_lower for word in ['открыт', 'новая', 'to do', 'open', 'new']):
            return 'status-open'
        elif any(word in status_lower for word in ['отклонен', 'rejected', 'declined']):
            return 'status-rejected'
        elif any(word in status_lower for word in ['отложен', 'отложено', 'отложена']):
            return 'status-other'
        else:
            return 'status-other'

    # Собираем HTML
    html_parts = []
    html_parts.append("<!doctype html><html><head><meta charset='utf-8'><title>Сравнение ДИТ ↔ Invaders</title>")
    html_parts.append(css)
    html_parts.append("</head><body><div class='container'><h1>Сравнение ДИТ ↔ Invaders</h1>")
    html_parts.append("<div class='controls'>")
    html_parts.append("<button class='btn' onclick='expandAllSwimlanes()'>Развернуть все</button>")
    html_parts.append("<button class='btn' onclick='collapseAllSwimlanes()'>Свернуть все</button>")
    html_parts.append("<button class='btn' onclick='showAll()'>Показать всё</button>")
    html_parts.append("<button class='btn' onclick='hideAll()'>Скрыть всё</button>")
    html_parts.append("<button class='btn' onclick=\"toggleClass('match')\">Toggle совпадения</button>")
    html_parts.append("<button class='btn' onclick=\"toggleClass('diff')\">Toggle разные спринты</button>")
    html_parts.append("<button class='btn' onclick=\"toggleClass('mos-only')\">Toggle только ДИТ</button>")
    html_parts.append("<button class='btn' onclick=\"toggleClass('inv-only')\">Toggle только Invaders</button>")
    
    # Кнопка экспорта в Excel
    html_parts.append("<button class='export-btn' onclick=\"window.location.href='comparison_report.xlsx'\">Скачать Excel отчет</button>")
    
    html_parts.append("<span class='small'>Фильтры работают визуально</span>")
    html_parts.append("</div>")
    
    # Добавляем фильтр по задаче
    html_parts.append("<div class='filter-row'>")
    html_parts.append("<div class='filter-label'>Фильтр по задаче:</div>")
    html_parts.append("<input type='text' id='taskFilter' class='filter-input' placeholder='Введите номер или название задачи (META-123, MT-456, или текст)' onkeypress='handleFilterKeyPress(event)'>")
    html_parts.append("<button class='filter-btn' onclick='filterByTask()'>Фильтровать</button>")
    html_parts.append("<button class='filter-clear' onclick='clearFilter()'>Очистить</button>")
    html_parts.append("</div>")
    
    # Добавляем фильтр по спринту
    html_parts.append("<div class='filter-row'>")
    html_parts.append("<div class='filter-label'>Фильтр по спринту:</div>")
    html_parts.append("<select id='sprintFilter' class='filter-select'>")
    html_parts.append("<option value=''>Все спринты</option>")
    for sp in sorted_sprints:
        html_parts.append(f"<option value='{html.escape(sp)}'>{html.escape(sp)}</option>")
    html_parts.append("</select>")
    html_parts.append("<button class='filter-btn' onclick='filterBySprint()'>Применить</button>")
    html_parts.append("<button class='filter-clear' onclick='clearSprintFilter()'>Очистить</button>")
    html_parts.append("<button class='filter-clear' onclick='clearAllFilters()'>Очистить все фильтры</button>")
    html_parts.append("</div>")
    
    # Секция информации об экспорте
    html_parts.append("<div class='export-section'>")
    html_parts.append("<strong>Доступен экспорт в Excel:</strong>")
    html_parts.append("<div class='export-info'>")
    html_parts.append(f"• Отчет содержит {len(categorized['match'])} совпадений, {len(categorized['diff_sprint'])} задач с разными спринтами<br>")
    html_parts.append(f"• Только в ДИТ: {len(categorized['mos_only'])} задач<br>")
    html_parts.append(f"• Только в Invaders: {len(categorized['inv_only'])} задач<br>")
    html_parts.append(f"• Задачи: {total_regular}, Баги: {total_bugs}<br>")
    html_parts.append("• Нажмите кнопку 'Скачать Excel отчет' для выгрузки полных данных")
    html_parts.append("</div>")
    html_parts.append("</div>")
    
    html_parts.append("<div class='legend'><b>Легенда:</b> <span style='background:#e6f6ea;padding:4px 8px;border-radius:4px;margin-left:8px'>совпадение (зелёный)</span> <span style='background:#fff8e0;padding:4px 8px;border-radius:4px;margin-left:8px'>разные спринты (жёлтый)</span> <span style='background:#ffe9e9;padding:4px 8px;border-radius:4px;margin-left:8px'>только ДИТ (красный)</span> <span style='background:#e8f1ff;padding:4px 8px;border-radius:4px;margin-left:8px'>только Invaders (синий)</span> <span class='bug-indicator'>Баг</span> <span class='status-ready' style='padding:2px 6px;border-radius:4px;margin-left:8px'>Готово</span> <span class='status-inprogress' style='padding:2px 6px;border-radius:4px;margin-left:8px'>В работе</span> <span class='status-open' style='padding:2px 6px;border-radius:4px;margin-left:8px'>Открыто</span></div>")

    # Свимлайн для обычных задач
    html_parts.append(f"<div id='tasks-swinlane' class='swimlane'>")
    html_parts.append(f"<div class='swimlane-header' onclick='toggleSwimlane(\"tasks-swinlane\")'>")
    html_parts.append(f"<span class='swimlane-title'>Задачи ({total_regular})</span>")
    html_parts.append(f"<span class='swimlane-count'>+</span>")
    html_parts.append(f"</div>")
    html_parts.append(f"<div class='swimlane-content'>")
    
    # Контейнер для таблицы с задачами
    html_parts.append("<div class='table-container'>")
    
    # Создаем оригинальную таблицу для задач
    html_parts.append("<table><thead><tr>")
    for sp in sorted_sprints:
        html_parts.append(f"<th colspan='2'>{html.escape(sp)}</th>")
    html_parts.append("</tr><tr>")
    for _ in sorted_sprints:
        html_parts.append("<th class='col-head'>ДИТ</th><th class='col-head'>Invaders</th>")
    html_parts.append("</tr></thead><tbody><tr>")

    # Body: For each sprint, render DIT column and Invaders column для НЕ-багов
    for sp in sorted_sprints:
        # DIT column для задач (не багов)
        html_parts.append("<td>")
        # matches where both sprints equal this sp и НЕ баг
        for it in categorized['match']:
            if it.get('mos_sprint') == sp and it.get('inv_sprint') == sp and not it.get('is_bug', False):
                status_class = get_status_class(it.get('mos_status'))
                html_parts.append("<div class='task match'>")
                html_parts.append(f"<div class='id'>{html.escape(str(it.get('mos_id') or ''))}")
                if it.get('mos_status') and it.get('mos_status') != 'Неизвестно':
                    html_parts.append(f"<span class='status {status_class}'>{html.escape(str(it.get('mos_status')))}</span>")
                html_parts.append("</div>")
                html_parts.append(f"<div class='title'><a href='{it.get('mos_url', '#')}' target='_blank'>{html.escape(str(it.get('mos_title') or it.get('inv_title') or ''))}</a></div>")
                html_parts.append("</div>")
        # diff_sprint where mos_sprint == sp и НЕ баг
        for it in categorized['diff_sprint']:
            if it.get('mos_sprint') == sp and not it.get('is_bug', False):
                status_class = get_status_class(it.get('mos_status'))
                html_parts.append("<div class='task diff'>")
                html_parts.append(f"<div class='id'>{html.escape(str(it.get('mos_id') or it.get('inv_id') or ''))}")
                if it.get('mos_status') and it.get('mos_status') != 'Неизвестно':
                    html_parts.append(f"<span class='status {status_class}'>{html.escape(str(it.get('mos_status')))}</span>")
                html_parts.append("</div>")
                if it.get('mos_url') != '#':
                    html_parts.append(f"<div class='title'>MOS: <a href='{it.get('mos_url')}' target='_blank'>{html.escape(str(it.get('mos_title') or ''))}</a><br/>INV: <a href='{it.get('inv_url')}' target='_blank'>{html.escape(str(it.get('inv_title') or ''))}</a></div>")
                else:
                    html_parts.append(f"<div class='title'>MOS: {html.escape(str(it.get('mos_title') or ''))}<br/>INV: {html.escape(str(it.get('inv_title') or ''))}</div>")
                html_parts.append("</div>")
        # mos_only и НЕ баг
        for it in categorized['mos_only']:
            if it.get('mos_sprint') == sp and not it.get('is_bug', False):
                status_class = get_status_class(it.get('mos_status'))
                html_parts.append("<div class='task mos-only'>")
                html_parts.append(f"<div class='id'>{html.escape(str(it.get('mos_id') or ''))}")
                if it.get('mos_status') and it.get('mos_status') != 'Неизвестно':
                    html_parts.append(f"<span class='status {status_class}'>{html.escape(str(it.get('mos_status')))}</span>")
                html_parts.append("</div>")
                if it.get('mos_url') != '#':
                    html_parts.append(f"<div class='title'><a href='{it.get('mos_url')}' target='_blank'>{html.escape(str(it.get('mos_title') or ''))}</a></div>")
                else:
                    html_parts.append(f"<div class='title'>{html.escape(str(it.get('mos_title') or ''))}</div>")
                html_parts.append("</div>")
        html_parts.append("</td>")

        # Invaders column для задач (не багов)
        html_parts.append("<td>")
        for it in categorized['match']:
            if it.get('inv_sprint') == sp and it.get('mos_sprint') == sp and not it.get('is_bug', False):
                status_class = get_status_class(it.get('inv_status'))
                html_parts.append("<div class='task match'>")
                html_parts.append(f"<div class='id'>{html.escape(str(it.get('inv_id') or ''))}")
                if it.get('inv_status') and it.get('inv_status') != 'Неизвестно':
                    html_parts.append(f"<span class='status {status_class}'>{html.escape(str(it.get('inv_status')))}</span>")
                html_parts.append("</div>")
                html_parts.append(f"<div class='title'><a href='{it.get('inv_url', '#')}' target='_blank'>{html.escape(str(it.get('inv_title') or it.get('mos_title') or ''))}</a></div>")
                html_parts.append("</div>")
        for it in categorized['diff_sprint']:
            if it.get('inv_sprint') == sp and not it.get('is_bug', False):
                status_class = get_status_class(it.get('inv_status'))
                html_parts.append("<div class='task diff'>")
                html_parts.append(f"<div class='id'>{html.escape(str(it.get('inv_id') or it.get('mos_id') or ''))}")
                if it.get('inv_status') and it.get('inv_status') != 'Неизвестно':
                    html_parts.append(f"<span class='status {status_class}'>{html.escape(str(it.get('inv_status')))}</span>")
                html_parts.append("</div>")
                if it.get('inv_url') != '#':
                    html_parts.append(f"<div class='title'>INV: <a href='{it.get('inv_url')}' target='_blank'>{html.escape(str(it.get('inv_title') or ''))}</a><br/>MOS: <a href='{it.get('mos_url')}' target='_blank'>{html.escape(str(it.get('mos_title') or ''))}</a></div>")
                else:
                    html_parts.append(f"<div class='title'>INV: {html.escape(str(it.get('inv_title') or ''))}<br/>MOS: {html.escape(str(it.get('mos_title') or ''))}</div>")
                html_parts.append("</div>")
        for it in categorized['inv_only']:
            if it.get('inv_sprint') == sp and not it.get('is_bug', False):
                status_class = get_status_class(it.get('inv_status'))
                html_parts.append("<div class='task inv-only'>")
                html_parts.append(f"<div class='id'>{html.escape(str(it.get('inv_id') or ''))}")
                if it.get('inv_status') and it.get('inv_status') != 'Неизвестно':
                    html_parts.append(f"<span class='status {status_class}'>{html.escape(str(it.get('inv_status')))}</span>")
                html_parts.append("</div>")
                if it.get('inv_url') != '#':
                    html_parts.append(f"<div class='title'><a href='{it.get('inv_url')}' target='_blank'>{html.escape(str(it.get('inv_title') or ''))}</a></div>")
                else:
                    html_parts.append(f"<div class='title'>{html.escape(str(it.get('inv_title') or ''))}</div>")
                html_parts.append("</div>")
        html_parts.append("</td>")

    html_parts.append("</tr></tbody></table>")
    html_parts.append("</div>")  # Закрываем table-container
    html_parts.append("</div>")  # Закрываем swimlane-content
    html_parts.append("</div>")  # Закрываем swimlane

    # Свимлайн для багов
    html_parts.append(f"<div id='bugs-swinlane' class='swimlane'>")
    html_parts.append(f"<div class='swimlane-header' onclick='toggleSwimlane(\"bugs-swinlane\")'>")
    html_parts.append(f"<span class='swimlane-title'>Баги ({total_bugs}) <span class='bug-indicator'>БАГ</span></span>")
    html_parts.append(f"<span class='swimlane-count'>+</span>")
    html_parts.append(f"</div>")
    html_parts.append(f"<div class='swimlane-content'>")
    
    # Контейнер для таблицы с багами
    html_parts.append("<div class='table-container'>")
    
    # Создаем оригинальную таблицу для багов
    html_parts.append("<table><thead><tr>")
    for sp in sorted_sprints:
        html_parts.append(f"<th colspan='2'>{html.escape(sp)}</th>")
    html_parts.append("</tr><tr>")
    for _ in sorted_sprints:
        html_parts.append("<th class='col-head'>ДИТ</th><th class='col-head'>Invaders</th>")
    html_parts.append("</tr></thead><tbody><tr>")

    # Body: For each sprint, render DIT column and Invaders column для багов
    for sp in sorted_sprints:
        # DIT column для багов
        html_parts.append("<td>")
        # matches where both sprints equal this sp и баг
        for it in categorized['match']:
            if it.get('mos_sprint') == sp and it.get('inv_sprint') == sp and it.get('is_bug', False):
                status_class = get_status_class(it.get('mos_status'))
                html_parts.append("<div class='task match bug-task'>")
                html_parts.append(f"<div class='id'>{html.escape(str(it.get('mos_id') or ''))} <span class='bug-indicator'>БАГ</span>")
                if it.get('mos_status') and it.get('mos_status') != 'Неизвестно':
                    html_parts.append(f"<span class='status {status_class}'>{html.escape(str(it.get('mos_status')))}</span>")
                html_parts.append("</div>")
                html_parts.append(f"<div class='title'><a href='{it.get('mos_url', '#')}' target='_blank'>{html.escape(str(it.get('mos_title') or it.get('inv_title') or ''))}</a></div>")
                html_parts.append("</div>")
        # diff_sprint where mos_sprint == sp и баг
        for it in categorized['diff_sprint']:
            if it.get('mos_sprint') == sp and it.get('is_bug', False):
                status_class = get_status_class(it.get('mos_status'))
                html_parts.append("<div class='task diff bug-task'>")
                html_parts.append(f"<div class='id'>{html.escape(str(it.get('mos_id') or it.get('inv_id') or ''))} <span class='bug-indicator'>БАГ</span>")
                if it.get('mos_status') and it.get('mos_status') != 'Неизвестно':
                    html_parts.append(f"<span class='status {status_class}'>{html.escape(str(it.get('mos_status')))}</span>")
                html_parts.append("</div>")
                if it.get('mos_url') != '#':
                    html_parts.append(f"<div class='title'>MOS: <a href='{it.get('mos_url')}' target='_blank'>{html.escape(str(it.get('mos_title') or ''))}</a><br/>INV: <a href='{it.get('inv_url')}' target='_blank'>{html.escape(str(it.get('inv_title') or ''))}</a></div>")
                else:
                    html_parts.append(f"<div class='title'>MOS: {html.escape(str(it.get('mos_title') or ''))}<br/>INV: {html.escape(str(it.get('inv_title') or ''))}</div>")
                html_parts.append("</div>")
        # mos_only и баг
        for it in categorized['mos_only']:
            if it.get('mos_sprint') == sp and it.get('is_bug', False):
                status_class = get_status_class(it.get('mos_status'))
                html_parts.append("<div class='task mos-only bug-task'>")
                html_parts.append(f"<div class='id'>{html.escape(str(it.get('mos_id') or ''))} <span class='bug-indicator'>БАГ</span>")
                if it.get('mos_status') and it.get('mos_status') != 'Неизвестно':
                    html_parts.append(f"<span class='status {status_class}'>{html.escape(str(it.get('mos_status')))}</span>")
                html_parts.append("</div>")
                if it.get('mos_url') != '#':
                    html_parts.append(f"<div class='title'><a href='{it.get('mos_url')}' target='_blank'>{html.escape(str(it.get('mos_title') or ''))}</a></div>")
                else:
                    html_parts.append(f"<div class='title'>{html.escape(str(it.get('mos_title') or ''))}</div>")
                html_parts.append("</div>")
        html_parts.append("</td>")

        # Invaders column для багов
        html_parts.append("<td>")
        for it in categorized['match']:
            if it.get('inv_sprint') == sp and it.get('mos_sprint') == sp and it.get('is_bug', False):
                status_class = get_status_class(it.get('inv_status'))
                html_parts.append("<div class='task match bug-task'>")
                html_parts.append(f"<div class='id'>{html.escape(str(it.get('inv_id') or ''))} <span class='bug-indicator'>БАГ</span>")
                if it.get('inv_status') and it.get('inv_status') != 'Неизвестно':
                    html_parts.append(f"<span class='status {status_class}'>{html.escape(str(it.get('inv_status')))}</span>")
                html_parts.append("</div>")
                html_parts.append(f"<div class='title'><a href='{it.get('inv_url', '#')}' target='_blank'>{html.escape(str(it.get('inv_title') or it.get('mos_title') or ''))}</a></div>")
                html_parts.append("</div>")
        for it in categorized['diff_sprint']:
            if it.get('inv_sprint') == sp and it.get('is_bug', False):
                status_class = get_status_class(it.get('inv_status'))
                html_parts.append("<div class='task diff bug-task'>")
                html_parts.append(f"<div class='id'>{html.escape(str(it.get('inv_id') or it.get('mos_id') or ''))} <span class='bug-indicator'>БАГ</span>")
                if it.get('inv_status') and it.get('inv_status') != 'Неизвестно':
                    html_parts.append(f"<span class='status {status_class}'>{html.escape(str(it.get('inv_status')))}</span>")
                html_parts.append("</div>")
                if it.get('inv_url') != '#':
                    html_parts.append(f"<div class='title'>INV: <a href='{it.get('inv_url')}' target='_blank'>{html.escape(str(it.get('inv_title') or ''))}</a><br/>MOS: <a href='{it.get('mos_url')}' target='_blank'>{html.escape(str(it.get('mos_title') or ''))}</a></div>")
                else:
                    html_parts.append(f"<div class='title'>INV: {html.escape(str(it.get('inv_title') or ''))}<br/>MOS: {html.escape(str(it.get('mos_title') or ''))}</div>")
                html_parts.append("</div>")
        for it in categorized['inv_only']:
            if it.get('inv_sprint') == sp and it.get('is_bug', False):
                status_class = get_status_class(it.get('inv_status'))
                html_parts.append("<div class='task inv-only bug-task'>")
                html_parts.append(f"<div class='id'>{html.escape(str(it.get('inv_id') or ''))} <span class='bug-indicator'>БАГ</span>")
                if it.get('inv_status') and it.get('inv_status') != 'Неизвестно':
                    html_parts.append(f"<span class='status {status_class}'>{html.escape(str(it.get('inv_status')))}</span>")
                html_parts.append("</div>")
                if it.get('inv_url') != '#':
                    html_parts.append(f"<div class='title'><a href='{it.get('inv_url')}' target='_blank'>{html.escape(str(it.get('inv_title') or ''))}</a></div>")
                else:
                    html_parts.append(f"<div class='title'>{html.escape(str(it.get('inv_title') or ''))}</div>")
                html_parts.append("</div>")
        html_parts.append("</td>")

    html_parts.append("</tr></tbody></table>")
    html_parts.append("</div>")  # Закрываем table-container
    html_parts.append("</div>")  # Закрываем swimlane-content
    html_parts.append("</div>")  # Закрываем swimlane
    
    html_parts.append(js)
    html_parts.append("</div></body></html>")

    out_html = "".join(html_parts)
    out_file.write_text(out_html, encoding="utf-8")
    print("Saved HTML:", str(out_file))

# -------------------------
# Main - с улучшенным поиском спринтов
# -------------------------
def main():
    base = Path(__file__).parent
    mos_path = base / MOS_NAME
    inv_path = base / INV_NAME
    out_path = base / OUT_NAME
    excel_path = base / EXCEL_NAME

    if not mos_path.exists():
        print("Файл Mos.csv не найден в папке со скриптом:", mos_path)
        return
    if not inv_path.exists():
        print("Файл Invaders.csv не найден в папке со скриптом:", inv_path)
        return

    mos_df = read_csv_guess(mos_path)
    inv_df = read_csv_guess(inv_path)
    
    print("=" * 80)
    print("Анализ файлов...")
    print(f"Колонки Mos.csv: {list(mos_df.columns)}")
    print(f"Колонки Invaders.csv: {list(inv_df.columns)}")
    print("=" * 80)

    # гарантируем необходимые колонки
    if 'Тема' not in mos_df.columns:
        mos_df['Тема'] = mos_df.iloc[:, 0].astype(str)
    if 'Ключ проблемы' not in mos_df.columns:
        mos_df['Ключ проблемы'] = mos_df['Тема'].apply(lambda x: extract_meta_key_from_text(x) or "")

    # Invaders: стандартизируем
    if 'Тема' not in inv_df.columns:
        # если нет такой колонки, попробуем первые колонки
        inv_df['Тема'] = inv_df.iloc[:, 0].astype(str)
    
    # Поиск колонки со спринтом в Invaders
    print("\nПоиск колонки со спринтом в Invaders...")
    sprint_col = None
    
    # Возможные названия колонки со спринтом
    possible_names = [
        'Релизный спринт', 'Release Sprint', 'Sprint', 'Спринт',
        'Пользовательское поле (Релизный спринт)',
        'Custom field (Release Sprint)'
    ]
    
    # Ищем точное совпадение
    for col in inv_df.columns:
        col_str = str(col).strip()
        if col_str in possible_names:
            sprint_col = col
            print(f"  ✓ Найдена колонка спринта: '{sprint_col}'")
            break
    
    # Если не нашли точное совпадение, ищем частичное
    if not sprint_col:
        for col in inv_df.columns:
            col_lower = str(col).lower()
            if any(name.lower() in col_lower for name in ['спринт', 'sprint', 'релиз']):
                sprint_col = col
                print(f"  ⚠️ Найдена похожая колонка: '{col}'")
                break
    
    # Если все еще не нашли, показываем первые значения из каждой колонки
    if not sprint_col:
        print("  ❗ Не найдена колонка со спринтом. Проверяем содержимое колонок...")
        for col in inv_df.columns[:5]:  # Проверяем первые 5 колонок
            sample_values = inv_df[col].dropna().head(3)
            if not sample_values.empty:
                print(f"    Колонка '{col}': {list(sample_values.values)}")
                # Проверяем, содержат ли значения слово "спринт"
                for val in sample_values:
                    if isinstance(val, str) and ('спринт' in val.lower() or 'sprint' in val.lower()):
                        sprint_col = col
                        print(f"  ✓ Возможно это колонка спринта: '{col}'")
                        break
            if sprint_col:
                break
    
    if sprint_col:
        inv_df['Пользовательское поле (Релизный спринт)'] = inv_df[sprint_col]
    else:
        print("  ❗ Не удалось найти колонку со спринтом. Используем 'Нет спринта'")
        inv_df['Пользовательское поле (Релизный спринт)'] = None

    # нормализация: canonical sprint и извлечение ключа из темы Invaders
    mos_df['Компоненты'] = mos_df.get('Компоненты', None)
    mos_df['sprint'] = mos_df['Компоненты'].apply(canonical_sprint)
    
    inv_df['Ключ проблемы'] = inv_df.get('Ключ проблемы')  # если уже есть, оставим
    # если ключа нет, попытаемся извлечь из Тема
    inv_df['maybe_key'] = inv_df['Тема'].apply(extract_inv_key_from_text)
    inv_df['Ключ проблемы'] = inv_df.apply(lambda r: (
        r['Ключ проблемы'] if r['Ключ проблемы'] and not (isinstance(r['Ключ проблемы'], float) and pd.isna(r['Ключ проблемы'])) 
        else r['maybe_key']
    ), axis=1)
    inv_df['sprint'] = inv_df['Пользовательское поле (Релизный спринт)'].apply(canonical_sprint)
    
    # Статистика
    print(f"\nСтатистика по спринтам:")
    print(f"  ДИТ: {mos_df['sprint'].nunique()} уникальных спринтов")
    print(f"  Invaders: {inv_df['sprint'].nunique()} уникальных спринтов")
    
    if inv_df['sprint'].nunique() == 1 and inv_df['sprint'].iloc[0] == "Нет спринта":
        print("\n⚠️ ВНИМАНИЕ: В файле Invaders не найдены спринты!")
        print("Возможные причины:")
        print("  1. Колонка со спринтом называется по-другому")
        print("  2. В данных нет информации о спринтах")
        print("  3. Формат данных отличается от ожидаемого")
        print("\nПроверьте CSV файл и убедитесь, что есть колонка с названием спринтов.")

    # Выполняем матчи
    print("\nВыполняем сопоставление задач...")
    matches, mos_used, inv_used = match_two_way(mos_df, inv_df)
    
    print(f"\nРезультаты сопоставления:")
    print(f"  Найдено совпадений: {len(matches)}")
    print(f"  Задействовано задач из ДИТ: {len(mos_used)}")
    print(f"  Задействовано задач из Invaders: {len(inv_used)}")
    
    categorized = categorize_and_prepare(mos_df, inv_df, matches, mos_used, inv_used)
    
    print(f"\nКатегоризация:")
    print(f"  Совпадения (один спринт): {len(categorized['match'])}")
    print(f"  Совпадения (разные спринты): {len(categorized['diff_sprint'])}")
    print(f"  Только в ДИТ: {len(categorized['mos_only'])}")
    print(f"  Только в Invaders: {len(categorized['inv_only'])}")
    
    # Статистика по статусам
    print(f"\nСтатистика по статусам:")
    status_stats = {}
    for cat_name, cat_list in categorized.items():
        for item in cat_list:
            for sys in ['mos', 'inv']:
                status_key = f'{sys}_status'
                if status_key in item:
                    status = item[status_key]
                    if status not in status_stats:
                        status_stats[status] = 0
                    status_stats[status] += 1
    
    for status, count in sorted(status_stats.items()):
        print(f"  {status}: {count}")

    # генерируем HTML
    print(f"\nГенерация HTML отчета...")
    generate_html(categorized, out_path, mos_df, inv_df)
    
    # экспортируем в Excel
    try:
        print(f"\nЭкспорт в Excel...")
        export_to_excel(categorized, excel_path, mos_df, inv_df)
        print(f"✓ Excel отчет создан: {excel_path}")
    except ImportError:
        print("\n❌ Для экспорта в Excel требуется библиотека openpyxl.")
        print("Установите её командой: pip install openpyxl")
    except Exception as e:
        print(f"\n❌ Ошибка при создании Excel файла: {e}")
        import traceback
        traceback.print_exc()
    
    print("\n" + "=" * 80)
    print("Обработка завершена!")
    print(f"HTML отчет: {out_path}")
    print(f"Excel отчет: {excel_path}")
    print("=" * 80)

if __name__ == "__main__":
    main()
